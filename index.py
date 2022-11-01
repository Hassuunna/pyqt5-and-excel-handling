import sys
from os import path
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUiType
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
import pandasModel

sourcePath = path.join(path.dirname(__file__),'design3.ui')

formClass, baseClass = loadUiType(sourcePath)

class MainApp(formClass, baseClass):
    start, end = 0, 0
    intervals = []
    df = pd.DataFrame()
    result_df = pd.DataFrame()
    final_result = pd.DataFrame()
    def __init__(self, parent=None):
        super(MainApp, self).__init__(parent)
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.Browse_BTN.clicked.connect(self.addItemsToTable)
        self.Calculate_BTN.clicked.connect(self.Calculate)
        self.Append_BTN.clicked.connect(self.Append)
        self.Save_BTN.clicked.connect(self.Save)

    def addItemsToTable(self):
        myFileExcel = QFileDialog.getOpenFileName(self, '', '',"Excel files (*.xlsx *.xls)")[0]
        if myFileExcel == '':
            QMessageBox.about(self, "Error", "Please Choose a file")
            return
        self.df = pd.read_excel(myFileExcel)
        if self.df.size == 0:
            QMessageBox.about(self, "Warning", "file is empty")
            return
        self.df.fillna('', inplace=True)
        self.tableWidget.setRowCount(self.df.shape[0])
        self.tableWidget.setColumnCount(self.df.shape[1])
        self.tableWidget.setHorizontalHeaderLabels(self.df.columns)

        for row in self.df.iterrows():
            row_index = row[0]
            values = row[1]
            for col_index, myValue in enumerate(values):
                if isinstance(myValue, (float, int)):
                    myValue = int(myValue)
                mytableItem = QTableWidgetItem(str(myValue))
                self.tableWidget.setItem(row_index, col_index, mytableItem)


    def Calculate(self):
        #make this dynamic done
        first_header = self.df.columns.tolist()[0]
        mod_df = self.df.set_index(first_header)
        #what if startValue and endValue are empty? done
        self.start = int(self.startValue.text()) if not self.startValue.text() == '' else int(mod_df.index[0])
        self.end = int(self.endValue.text()) if not self.endValue.text() == '' else int(mod_df.index[-1])
        if self.end < self.start: 
            temp = self.end
            self.end = self.start
            self.start = temp
        interval = mod_df.loc[self.start: self.end]
        print(interval)
        # round done
        self.result_df = interval.aggregate(['min', 'max', 'mean']).astype(int)
        model = pandasModel(self.result_df)
        self.tableView.setModel(model)


    def Append(self):
        # turn interval into string
        self.intervals.append(str(self.start) + " - " + str(self.end))
        # define new empty row to insert
        new_row = pd.Series({})
        # concat the old existing data then empty row then new data
        self.final_result = pd.concat([self.final_result, new_row.to_frame().T, self.result_df])


    def Save (self):
        #append start and end done
        #add merged big cell done
        #don't overwrite done
        # get file name
        excelFileName = QFileDialog.getSaveFileName(self, 'Save as', '', 'Excel (*.xlsx)')[0]
        # startcol = 1 to leave space for intervals
        self.final_result.to_excel(excelFileName, startcol=1)

        #with pd.ExcelWriter(myNewName, mode="a", if_sheet_exists='overlay') as writer:
        # to open the excel sheet and if it has macros
        wb = load_workbook(excelFileName, read_only=False)
        
        # get sheetname from the file
        sheetname = wb['Sheet1']
        
        #count existing rows in sheet
        begin = sheetname.max_row + 2 - len(self.intervals) * 4
        
        # write to row ,col explicitly, this type of writing is useful to
        # write something in loops
        for interval in self.intervals:
            #merge 3 cells corresponding to min max mean
            sheetname.merge_cells('A'+str(begin)+':A'+str(begin+2))
            # remove zero from column B in index5.2
            sheetname['B' + str(begin-1)] = ""
            # write interval string
            sheetname['A' + str(begin)] = interval
            # align cell text in center
            cell = sheetname.cell(begin, 1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # increment by 4 to write the next interval
            begin+=4
        
        # save it as the original file
        wb.save(excelFileName)


def main():
    myApp = QApplication(sys.argv)
    myWindow = MainApp()
    myWindow.show()
    myApp.exec_()

if __name__ == '__main__':
    main()